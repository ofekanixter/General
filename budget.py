import os
import openpyxl as xl
import re

import pandas as pd

#for detect if a string is english word
english_pattern = r'^[A-Za-z\s]+$'

#files constant path
EXCEL_FOLDER="C:/Users/SmadarENB3/Desktop/ofek/programing/python/budget/excel"
BANK_PATH=EXCEL_FOLDER+"/budget_bank.xlsx"
KIBBUTZ_PATH=EXCEL_FOLDER+"/budget_kibbutz.xlsx"
CREDIT_PATH=EXCEL_FOLDER+"/budget_credit_"
SUMMARY_PATH=EXCEL_FOLDER+"/summary_budget_2023.xlsx"
CATEGORIES_PATH=EXCEL_FOLDER+"/categories.txt"

#categories list as each is (cat,cordinate), and a list to print
CAT_LIST=[("להתעלם",-1),("אוכל","C9"),("לבית","C17"),("בריאות","C25"),("בידור","C33"),
          ("קבועות","G9"),("כרמי","G18"),("קיבוץ","G26"),("רכב","G33"),("הכנסה","D4")]
CAT_LIST_FOR_CHOOSE=[s[::-1] for s in ["0להתעלם","אוכל","לבית","בריאות","4 בידור","קבועות","6 כרמי","קיבוץ","רכב 8","הכנסה 9"]]
COMMENTS="B40"

MONTH_DIC={"01":("ינואר","jan"),"02":("פברואר","feb"),"03":("מרץ","mar"),"04":("אפריל","apr"),"05":("מאי","may"),"06":("יוני","jun"),
           "07":("יולי",'jul'),"08":("אוגוסט","aug"),"09":("ספטמבר","sep"),"10":("אוקטובר","oct"),"11":("נובמבר","nov"),"12":("דצמבר","dec")}

#constants cordinate for start and end of the files
KIBB_ST='A2'
KIBB_END='D130'
CREDIT_ST='B2'
CREDIT_END='D60'
BANK_ST='A2'
BANK_END='E158'
SUMM_ROW="111"
SUMM_LIST= ['פרוט החודש','תאריך','רני','אופק','מיוחדים']
MONTH_NAME="יוני"

#Dollar rate 
DOLLAR=3.6
EURO=3.8

#s string if english ignore if hebrew reverse for printing
def hebrew_fix(s):
    s=str(s)
    if(s is None):
        return "None"
    if re.search(english_pattern, s.replace("#","").replace("/","")):
        return s
    if s.endswith("\n"):
        s=s[:-1]
        return s[::-1]+"\n"
    return s[::-1]

#just print after use hebrew_fix func on "s"
def print_hebrew(s):
    print(hebrew_fix(s))

#take a bill title name,, and if is one of the next patterns return the simpler pattern  
def find_pattern(name):
    pattern_list=[('ALIBABA', 'ALIBABA'), ('ALIEXPRESS', 'ALIEXPRESS'), ('AMZN MKTP', 'AMAZON'), ('AMAZON', 'AMAZON'),('APPLE', 'APPLE'),
                   ('ELAL', 'ELAL'), ('PAYPAL', 'PAYPAL'), ('בנהפ BIT', 'BIT'), ('סופר פארם', 'סופר פארם')]
    for pattern,short in pattern_list:
        if name.find(pattern)!=-1:
            return short
    if(name.find("AMZN MKTP")!=-1):
        return "AMAZON"
    if(name.find("PAYPAL")!=-1):
        return "PAYPAL"
    if(name.find("בנהפ BIT")!=-1):
        return "BIT"
    if(name.find("ALIEXPRESS")!=-1):
        return "ALI EXPRESS"
    if(name.find("ALIBABA")!=-1):
        return "ALIBABA"
    if(name.find("ELAL")!=-1):
        return "ELAL"
    if(name.find("APPLE")!=-1):
        return "APPLE"
    if(name.find("סופר פארם")!=-1):
        return "סופר פארם"
    
    return name

#take the Workbook and do finall steps as
#
def end_summary(wb,month_name="יוני"):
    sheet=wb[month_name]
    sheet['A2']="תקציב חודש- "+month_name
    income=input(hebrew_fix("יש הכנסות?\n"))
    if income=="D":
        return sheet
    if income=="":
        sheet['D4']=int(input(hebrew_fix("הכנסה 1\n")))
        sheet['D5']=int(input(hebrew_fix("הכנסה 2\n")))
    return sheet
#take an input and check if it valid else ask again
#for numeric=True , check if < limit
def input_check(inp,limit,numeric):
    if numeric:
        if inp.isnumeric():
            inp=int(inp)
            if inp<limit:
                return inp
            else:
                print("Wrong, you gave {} need to be <{}".format(inp,limit))
                return input_check(input("try agian\n"),limit,numeric)
        else:
            print("Wrong, you gave {} need to be a number".format(inp))
            return input_check(input("try agian\n"),limit,numeric)
    print("why")
    return inp

#for credit step, take Excel cells and 
#transfer them to a dict as {cat:[(bill_name,ammount),etc]} which return
def credit_cells_to_dic(cells,title):
    cat_dic=dict()
    end_shekel=False
    for B,C,D in cells:#column B and C at the file
        if B.value == "מטח" or B.value == None:#second part of the file(dollar bills)
            print(f'b= {B.value} and indexes are {B}')
            end_shekel=True
            continue
        #take the name, and ask user what cat it belong
        name=find_pattern(B.value.replace('\u200e',""))
        ammount = C.value
        #for foriegn coin
        if end_shekel:
            coin= EURO if D.value.find("אירו")!=1 else DOLLAR
            ammount = C.value*coin
        inp=input(hebrew_fix(name)+" , {} ILS - what cat? \n{}\n".format(ammount,CAT_LIST_FOR_CHOOSE))
        index=input_check(inp,len(CAT_LIST),True)
        cat=CAT_LIST[index][0]
        print("you choose cat -"+hebrew_fix(cat))
        #enter to the right cat the (name,ammount)  
        if cat in cat_dic:
            cat_dic[cat].append((name,ammount))
        else:
            cat_dic[cat]=[(name,ammount)]
    return cat_dic

#take the date, and check if it belong to the correct month(sheet title)
#or the next month beacuse some bills and income is for the last month       
def relevant_date(date,title):
    month=date.month
    next_month=month-1 if month>1 else 12
    month="0"+str(month) if month<10 else str(month)
    next_month="0"+str(next_month) if next_month<10 else str(next_month)
    if(MONTH_DIC[month][0]==title or MONTH_DIC[next_month][0]==title ):
        if (MONTH_DIC[next_month][0]==title and date.day>5):
            return False
        return True
    return False
#for credit step, take Excel cells and 
#transfer them to a dict as {cat:[(bill_name,ammount),etc]} which return
def bank_cells_to_dic(cells,title):
    month=0
    cat_dic=dict()
    for A,B,C,D,E in cells:#columns
        if E.value is None:
            break
        if not relevant_date(E.value,title):
            continue
        #take the name, and ask user what cat it belong
        name=find_pattern(C.value.replace('\u200e',""))
        ammount = (-1*A.value) if isinstance(B.value,str) else B.value
        inp=input(hebrew_fix(name)+" , {} ILS - what cat? \nDate is {}\n{}\n".format(ammount,E.value,CAT_LIST_FOR_CHOOSE))
        index=input_check(inp,len(CAT_LIST),True)
        cat=CAT_LIST[index][0]
        print("you choose cat -"+hebrew_fix(cat))

        #enter to the right cat the (name,ammount)  
        if cat in cat_dic:
            cat_dic[cat].append((name,ammount))
        else:
            cat_dic[cat]=[(name,ammount)]
    return cat_dic

#for kibbutz step, take Excel cells and 
#transfer them to a dict as {cat:[(bill_name,ammount),etc]}
def kibb_cells_to_dic(cells,title):
    cat_dic=dict()
    for A,B,C,D in cells:#columnA,B,C and D at the file
        if D.value is None:
            break
        month=D.value[3:5]#checking that the bill is in the right month
        if(MONTH_DIC[month][0]!=title):
            continue
        name=C.value.replace('\u200e',"") 
        inp=input(hebrew_fix(name)+"- what cat? \n{}\n".format(CAT_LIST_FOR_CHOOSE))
        index=input_check(inp,len(CAT_LIST),True)
        cat=CAT_LIST[index][0]
        print("you choose cat -"+hebrew_fix(cat))
        if(B.value is None):
            ammount = A.value*-1
        else:
            ammount = B.value
        if cat in cat_dic:
            cat_dic[cat].append((name,ammount))
        else:
            cat_dic[cat]=[(name,ammount)]
    return cat_dic

#take a bill name,category and sub categeroy and insert to a dict
# at this tamplate {bill name:(cat,sub_cat)}
def insert_name_to_dict2(dic,name,cat ,sub_cat):
    if name in dic:
        return dic
    else:
        dic[name]=(cat,sub_cat)
    return dic

#take a bill name,category and sub categeroy and insert to a dict
# at this tamplate {cat:{sub_cat:{bill name}}
def insert_name_to_dict(dic,name,cat ,sub_cat):
    if cat in dic:
        if sub_cat in dic[cat]:
            dic[cat][sub_cat].add(name)
        else:
            dic[cat][sub_cat]={name}
    else:
        dic[cat]={sub_cat:{name}}
    return dic

#take the  cat_dict {cat:[(bill_name,ammount),etc]}
#and ask from user the right sub_cat
#and write at worksheet[cordinate(the right sub_cat)] the formula
# "=ammount1+ammount2+etc.." 
def write_dic_wb(cat_dic,sum_sheet,memory_dict,names_dict):
    TOTAL_AMMOUNT=0
    INCOME_AMMOUNT=0
    for cat_cell in CAT_LIST:
        cat= cat_cell[0]
        cell= cat_cell[1]
        if cat=="להתעלם":
            continue
        if cat in cat_dic:#else not need towrite anything
            for name,ammount in cat_dic[cat]:#for each biil name and its ammount at the curr cat
                #extract the col and row digits
                col=cell[0]
                col_digit=ord(col)-64
                row=int(cell[1:])
                #extract all the sub_cat in the curr cat
                #its 2 col to the right, and end with "סך הכל"
                #Then take get from the user the right sub_cat
                sub_cat=[" " if sum_sheet.cell(column=col_digit-2,row=i)==None else sum_sheet.cell(column=col_digit-2,row=i).value for i in range(row,row+7)] 
                sub_cat_to_print=[hebrew_fix(s) if s!="סך הכל" else "end" for s in sub_cat]
                limit=sub_cat_to_print.index("end")
                msg="{} cost {} ILS- what sub cat? index <{} {} \n".format(hebrew_fix(name),ammount,str(limit),str(sub_cat_to_print))
                inp=input(msg)
                index=input_check(inp,limit,True)
                #set the right cord for writing the ammount
                #and enter to the memory dict
                cord=col+str(row+index)
                memory_dict=insert_name_to_dict(memory_dict,name,cat ,sub_cat[index])
                names_dict=insert_name_to_dict2(names_dict,name,cat ,sub_cat[index])
                if sum_sheet[cord].value == None :
                    if ammount<0:
                        sum_sheet[cord].value ="="+str(-1*ammount)
                    else:
                        sum_sheet[cord].value ="="+str(ammount)
                else:
                    if ammount<0:
                        sum_sheet[cord].value +=str(ammount)
                    else:
                        sum_sheet[cord].value +="+"+str(ammount)
                #sum the total ammount per part(kibbutz,credit..)
                #ammount<0 meaning its income
                if ammount>0:    
                    TOTAL_AMMOUNT+=ammount
                else:
                    INCOME_AMMOUNT+=-1*ammount
                print("you choose sub cat -"+hebrew_fix(sub_cat[index]))

    #return the sheet after written, both update dicts and the total/income ammount
    return sum_sheet,memory_dict,names_dict,TOTAL_AMMOUNT,INCOME_AMMOUNT

#the kibbutz part flow,input the summary_sheet and the memory dicts
#load and active the kibbutz wb->extract relevent cells range->
# build a cat_dicr-> write to the sum_sheet and return the result
def kibbutz_sum(sum_sheet,memory_dict,names_dict):
    wb = xl.load_workbook(KIBBUTZ_PATH)
    kib_sheet = wb.active
    cells = kib_sheet[KIBB_ST:KIBB_END]
    cat_dic=kibb_cells_to_dic(cells,sum_sheet.title)
    return write_dic_wb(cat_dic,sum_sheet,memory_dict,names_dict)

#the credit card part flow,same as kibbutz
def credit_sum(sum_sheet,memory_dict,names_dict,month_number):
    suffix=MONTH_DIC[month_number][1]+".xlsx"
    os.startfile(CREDIT_PATH+suffix)
    wb = xl.load_workbook(CREDIT_PATH+suffix)
    sheet = wb.active
    cells = sheet[CREDIT_ST:CREDIT_END]
    cat_dic=credit_cells_to_dic(cells,sum_sheet.title)
    return write_dic_wb(cat_dic,sum_sheet,memory_dict,names_dict)

#the bank card part flow,same as kibbutz
def bank_sum(sum_sheet,memory_dict,names_dict):
    wb = xl.load_workbook(BANK_PATH)
    sheet = wb.active
    cells = sheet[BANK_ST:BANK_END]
    cat_dic=bank_cells_to_dic(cells,sum_sheet.title)
    return write_dic_wb(cat_dic,sum_sheet,memory_dict,names_dict)

#take a text file and extract it to a dict 
#format like {name:(cat,sub_cat)}
def file_to_dic2(path):
    cat_memory_dic=dict()
    if os.stat(path).st_size==0:#empty file
        return cat_memory_dic
    filee = open(path,'r')
    text=filee.read()
    filee.close
    cat_memory_dic=eval(text)#from string to dict
    return cat_memory_dic

#take a text file and extract it to a dict 
#format like {cat:{sub_cat:{name,name2,..}}}
def file_to_dic(path):
    cat_memory_dic=dict()
    if os.stat(path).st_size==0:#empty file
        return cat_memory_dic
    filee = open(path,'r')
    more=True
    new_cat=False
    new_sub=False
    line_cat=""
    while more:#end file
        #for new cat dont need to read new line, already read in the last iter
        if not new_cat:
            line=filee.readline()
            if line=="":
                return dict()
        else:
            line=line_cat
        if not line:#meaning EOF
            break
        cat=line[:-2]
        cat_set=dict()
        new_sub=False
        while(True):
            if not new_sub:
                sub_cat=filee.readline()[:-2].strip('\t')
            else:
                sub_cat=line_sub[:-2].strip('\t')
            #take down the \n and the \t and split with ',' to get all names
            names=set(filee.readline()[:-1].strip('\t').split(','))
            cat_set[sub_cat]=names
            line=filee.readline()
            if not line:#EOF
                more=False
                break
            elif line[-2]==':':#meaning new cat
                new_cat=True
                line_cat=line
                break
            else:#meaning new sub_cat
                new_sub=True
                line_sub=line
        cat_memory_dic[cat]=cat_set
    filee.close
    return cat_memory_dic

#from dic to file, one line 
def dic_to_file2(dic,path):
    filee=open(path,'w')
    filee.write(str(dic))
    filee.close

#from dic to file like:
#cat 
#   sub cat
#       name1,name2...
def dic_to_file(dic,path):
    text=""
    hash_dic=dict()
    for cat in dic:
        text+=cat+":\n"
        for sub_cat in dic[cat]:
            text+="\t"+sub_cat+"-\n"
            items=""
            for item in dic[cat][sub_cat]:
                items+=item+","
                hash_dic[item]=(cat,sub_cat)
            text+="\t\t"+items[:-1]+"\n"
    filee=open(path,'w')
    filee.write(text)
    filee.close

#before start summary
#open dics, load wb, get the month
def pre(path):
    print("start pre")
    print("#########")
    memory_dict=file_to_dic(path)    
    names_dict=file_to_dic2(path[:-4]+"_names.txt")
    wb = xl.load_workbook(SUMMARY_PATH)
    wb.active
    month_number=input("month number(01,02,03):\n")
    name=MONTH_DIC[month_number][0]
    print(hebrew_fix(name))
    sheet=wb[name]
    print("finish pre")
    print("#########")
    return wb,sheet,memory_dict,names_dict,month_number
    
def summary(KIBUTTZ_CREDITS_BANK,month_name):
    wb = xl.load_workbook(SUMMARY_PATH)
    wb.active
    sheet = wb['כללי']
    outputs=list()
    for box in SUMM_LIST:
        if box=='תאריך':
            outputs.append(month_name)
            continue
        outputs.append(input(f'enter input for: {hebrew_fix(box)}\n'))
    outputs.extend(KIBUTTZ_CREDITS_BANK)
    outputs[-1]+=300
    OK=True
    while(OK):
        row=input("enter row number\n")
        if row=='':
            row=SUMM_ROW
        print(f'the list is:\n{outputs}')
        OK=input(f"row is {row}\nif OK put Y else enter\n")!='Y'
        if not OK :
            break
        index=int(input("what index to changed?\n"))
        while not isinstance(index,int):
            index=input(f"we need a digit please\n")
        outputs[index]=input("what is the right content\n")
    print("we are OK!")
    #save outputs in the right place
    j=0
    for i in range(len(outputs)):
        if i==4:
            j=1
        #print(f'colrow={chr(65+i+j)+SUMM_ROW} and output is :{outputs[i]}')
        sheet[chr(65+i+j)+row]=outputs[i]
    #save
    try:
        input("save?")
        wb.save(SUMMARY_PATH)
    except PermissionError:
        input("close excel file")
        wb.save(SUMMARY_PATH)
    print("The wb saved")

def main():
    summary((1,2,3),month_name='פברואר')
    if input("open files? enter for yes\n")=="":
        os.startfile(SUMMARY_PATH)
        os.startfile(BANK_PATH)
        os.startfile(KIBBUTZ_PATH)
    #memory_dict=dict()
    #print(file_to_dic(CATEGORIES_PATH))
    # extract memories dict, open summary wb and the right sheet
    wb,sum_sheet,memory_dict,names_dict,month_number=pre(CATEGORIES_PATH)    

    #start the kibbutz summary, sheet and dicts are filled also the total ammount
    print("start kibbutz")
    print("#############")   
    sum_sheet,memory_dict,names_dict,TOTAL_AMMOUNT_KIBUTTZ,a=kibbutz_sum(sum_sheet,memory_dict,names_dict)
    print("start credits")
    print("#############") 
    #start the credit summary, sheet and dicts are filled also the total ammount
    sum_sheet,memory_dict,names_dict,TOTAL_AMMOUNT_CREDITS,b=credit_sum(sum_sheet,memory_dict,names_dict,month_number)
    print("start bank")
    print("#############")
    #start the credit summary, sheet and dicts are filled also the total ammount
    sum_sheet,memory_dict,names_dict,TOTAL_AMMOUNT_BANK,INCOME_BANK=bank_sum(sum_sheet,memory_dict,names_dict)
    print("total ammount in the bank is {} and income is{}".format(TOTAL_AMMOUNT_BANK,INCOME_BANK))
    
    print("total ammount in the kibbutz is {}".format(int(TOTAL_AMMOUNT_KIBUTTZ)))
    print("total ammount in the credits is {}".format(int(TOTAL_AMMOUNT_CREDITS)))
    print("total ammount in the bank is {} and income is{}".format(TOTAL_AMMOUNT_BANK,INCOME_BANK))

    #try save wb, usually if failed it because that the excel file is open
    try:
        input("save?")
        wb.save(SUMMARY_PATH)
    except PermissionError:
        input("close excel file")
        wb.save(SUMMARY_PATH)
    print("The wb saved")
    #open for checks the file
    os.startfile(SUMMARY_PATH)
    #save meta month budget data
    summary((TOTAL_AMMOUNT_KIBUTTZ,TOTAL_AMMOUNT_CREDITS,TOTAL_AMMOUNT_BANK),month_name=MONTH_DIC[month_number][0])
    #open for checks the file
    os.startfile(SUMMARY_PATH)
    #save the updated dict to the text files
    print("save updates dicts")
    input("save?")
    dic_to_file(memory_dict,CATEGORIES_PATH)
    dic_to_file2(names_dict,CATEGORIES_PATH[:-4]+"_names.txt")


if __name__ == "__main__":
    main()
